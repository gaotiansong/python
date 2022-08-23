import requests
import json
import openpyxl

def postdata(Cookie,billCode,packageNo):
    headers={
        "cookie": Cookie,
        'Content-Type': 'application/json',
    }
    data={
        "operationType": 4,
        "billCode": billCode,
        "packageNo": packageNo,
        "weight": None,
        "businessId": None,
        "businessCode": None,
        "businessName": None,
        "preOrNextSiteCode": None,
        "preOrNextSiteName": None,
        "preOrNextSiteId": None,
        "goodsType": "1",
        "transportType": "2"
    }
    url="https://xpom.zt-express.com/operation/businessScanOperate"
    reg=requests.post(url, headers=headers,data=json.dumps(data))
    print("提交完毕:",reg.text)
    #{"status":true,"message":"操作成功","result":{"siteId":1065941,"siteCode":"539029","siteName":"临沂河东经开区","operationType":4,"billCode":"78280031222394","packageNo":"AZ94982921","weight":"0.00","businessId":null,"businessCode":null,"businessName":null,"preOrNextSiteCode":null,"preOrNextSiteName":null,"preOrNextSiteId":null,"goodsType":1,"transportType":2,"operateName":"冯彭蕾","operateId":5896651,"operateCode":"539029.10004","destination":"济宁中转部","destinationCode":"53739","scanDate":"20220822225142"},"statusCode":"SYS000"}
    #{"status":false,"message":"请求类型错误:application/json1;charset=UTF-8 不支持","result":null,"statusCode":"PARAM_ERROR"}
    return (json.loads(reg.text))["status"]

def red_cookie(file):
    with open(file,"r") as f:
        reads=f.read()
        return reads.strip()

if __name__ == '__main__':
    p1=input("输入cookie文件路径:")
    #/Users/gaotiansong/Desktop/cookie.txt
    p1=p1.strip()
    while True:
        try:
            c = red_cookie(p1)
            cookie=c
            break
        except Exception as e:
            print(e)
            p1 = input("重新输入cookie文件路径:")
            p1 = p1.strip()
    p2 = input("请输入表格路径:")
    p2 = p2.strip()
    while p2[-5:]!=r".xlsx":
        p2=input("重新输入表格路径:")
    #path = "/Users/gaotiansong/Desktop/副本测试数据.xlsx"
    path=p2
    print("cookie==",cookie)
    print("path==",path)

    book = openpyxl.load_workbook(path)
    sheet = book.active
    for row in range(sheet.max_row):
        v1 = sheet.cell(row=row + 1, column=1).value
        v2 = sheet.cell(row=row + 1, column=2).value
        if len(v1)<10:
            print("跳过==",v1, v2)
            continue
        print("数据==", v1, v2)
        try:
            status = postdata(cookie, v1, v2)
        except Exception as e:
            print(e)
            break
        if status:
            print("提交成功==",v1, v2)
        else:
            print("提交失败==",v1, v2)

    print("完成！")
