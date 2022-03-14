# coding=utf-8
from openpyxl import load_workbook


class ReadExec:
    def __init__(self, path):
        self.path = path
        print("开始加载:", self.path)
        self.wb = load_workbook(self.path)
        print("成功加载")
        self.Template = self.wb["Template"]
        print("获取my_sheet")
        for c in range(1, self.Template.max_column):  # 查找每一列 找出对应的字段所在的列
            if self.Template.cell(3, c).value == "item_name":
                self.c_name = c
            if self.Template.cell(3, c).value == "generic_keywords":
                self.c_keywords = c
            if self.Template.cell(3, c).value == "size_name":
                self.c_size_name = c
            if self.Template.cell(3, c).value == "main_image_url":
                self.c_image1 = c
            if self.Template.cell(3, c).value == "other_image_url1":
                self.c_image2 = c
            if self.Template.cell(3, c).value == "other_image_url2":
                self.c_image3 = c
        self.max_r = self.Template.max_row

    def sava_full(self, path):
        self.wb.save(path)


class ReadKeyWords:
    def __init__(self, path):
        self.myWB = None  # 核心词
        self.path = path
        self.wb = load_workbook(self.path)

    def get_words(self, sheet):
        # 获取核心词
        self.myWB = self.wb[sheet]
        words = []
        for row in range(1, self.myWB.max_row):
            word = self.myWB.cell(row, 1).value
            words.append(word)
        words = list(set(words))
        return words

    def get_rows(self, sheet):
        dbs = self.wb[sheet]
        return dbs
