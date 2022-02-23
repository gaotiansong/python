# coding=utf-8
from openpyxl import load_workbook,Workbook


class ReadExec:
    path = ""
    try:
        wb = load_workbook(path)
        my_sheet = wb["Template"]
    except Exception as e:
        wb = Workbook()
        my_sheet = wb.active
        my_sheet.title = "Template"
        ##print(e)
    # 确定标题和图片所在列
    c_name = 0
    c_image1 = 0
    c_image2 = 0
    c_image3 = 0
    max_r = my_sheet.max_row

    def open_excl(self, path):
        self.path = path
        self.wb = load_workbook(path)
        self.my_sheet = self.wb["Template"]
        for c in range(1, self.my_sheet.max_column):
            if self.my_sheet.cell(3, c).value == "item_name":
                self.c_name = c
            if self.my_sheet.cell(3, c).value == "main_image_url":
                self.c_image1 = c
            if self.my_sheet.cell(3, c).value == "other_image_url1":
                self.c_image2 = c
            if self.my_sheet.cell(3, c).value == "other_image_url2":
                self.c_image3 = c
        self.max_r = self.my_sheet.max_row

    def get_image(self, n):
        ##print("go get_image")
        n = n + 3  # 从第四行开始
        url = self.my_sheet.cell(n, self.c_image1).value
        name = self.my_sheet.cell(n, self.c_name).value
        ##print("url=", url)
        ##print("ok get_image")
        return url, name

    def sava_f(self, n, tx):
        n = n + 3  # 从第四行开始
        # 给格子赋值
        self.my_sheet.cell(n, self.c_name).value = tx
        ##print("保存=", tx)
        self.wb.save(self.path)
