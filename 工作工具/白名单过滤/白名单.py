#coding=utf-8
import pymysql
import re
from openpyxl import load_workbook


def find_cursor():
    # 打开数据库连接
    db = pymysql.connect(host='',
                        user='',
                        password='',
                        database='')

    # 使用 cursor() 方法创建一个游标对象 cursor
    cursor = db.cursor()
    return cursor,db

def find_datas(tablename):
    #从数据库中获取关键词列表
    datas=[]
    cursor,db=find_cursor()
    sql="select * from {tablename};".format(tablename=tablename)
    try:
        cursor.execute(sql)
        db.commit()
    except Exception as e:
        print(e)
        db.rollback()
    results = cursor.fetchall()
    db.close()
    for row in results:
        datas.append(row[1].lower())
    datas=list(set(datas))
    return datas


def write_to_excel(path,data,newsheet):
    #实例化一个workbook对象
    workbook = load_workbook(path)
    #激活一个sheet 如果不存在就创建它然后激活
    try:
        sheet1 = workbook[newsheet]
    except:
         #创建一个标签
        workbook.create_sheet(newsheet)
        sheet1 = workbook[newsheet]

    #开始遍历数组
    for row_index, row_item in enumerate(data):
        for col_index, col_item in enumerate(row_item):
            # 写入
            sheet1.cell(row=row_index+1,column= col_index+1,value=col_item)

    workbook.save(path)
    print('处理完毕')

#用来把混编字符串列表变成单字字符串列表
def change_ls(old_ls):
    s=" ".join(old_ls)
    s=s.lower() #转化为小写
    new_ls=s.split(" ")
    return new_ls

if __name__=="__main__":
    print("加载白名单关键词")
    #okkeywords=["church","Collection","Church"]
    stopkeywords=find_datas(tablename="stopword")
    tablename="okword"
    okkeywords=find_datas(tablename)
    print("成功加载okword")
    #要处理的文件
    #f_path=r"/Users/gaotiansong/Desktop/6187b3360cf24ef0eebbd7ad.xlsx"
    f_path=input("拉入你要进行白名单处理的表格:")
    f_path=re.sub(r'"',"",f_path)
    f_path=f_path.strip()

    wb = load_workbook(f_path)
    #选择标签
    mysheet=wb["Template"]
    new_data=[]
    for r in range(mysheet.max_row):
        row_s=[]
        row=[]
        for c in range(mysheet.max_column):
            #row1是每一个格子
            row1=mysheet.cell(r+1,c+1).value
            #row是每一行
            if row1==None:
                row1=""
            row.append(str(row1))
        #检测row中是否有违禁词
        print("row=",row)
        for stop in stopkeywords:
            #取出每一个违禁词
            if stop.lower() in change_ls(row): 
                #stop是违禁词 装入违禁词容器 row_s中
                row_s.append(stop)
        #获取包含包含违禁词的字符串p_word
        p_word=row[-1]
        mo=re.compile("\[.*?\]")
        #提取其中的违禁词
        ss=re.findall(mo,str(p_word))
        for s in ss:
            #取出每一个违禁词
            s=s[1:-1]
            for s1 in s.split(","): #把关键词变成列表['abc','cdg']
                #判断s1是否在白名单中，在则跳过，不在则标记为违禁词
                if s1.lower() not in okkeywords:
                    row_s.append(s1)
        rowss="|".join(list(set(row_s)))
        new_data.append(row+[rowss])

    #新标签名
    newsheet="New_Template"
    write_to_excel(f_path,new_data,newsheet)
    _=input("处理完毕 按任意键退出")
