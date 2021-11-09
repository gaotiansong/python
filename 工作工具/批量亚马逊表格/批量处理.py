#coding=utf-8
import csv
from openpyxl import load_workbook
from  openpyxl import  Workbook 
from openpyxl.cell import cell

old_csv=r"C:\Users\ZG\Downloads\211109待修改\All+Listings+Report+11-08-2021-Greanda30-ca.txt update.csv"

#实例化一个对象
work = Workbook()
#激活标签
sheet=work.active
#更改标签名
sheet.title="Template"
#打开模板
wb = load_workbook(r"C:\Users\ZG\Downloads\211109待修改\MUG.xlsm")
mysheet=wb["Template"]

with open(old_csv,"r") as f:
    rows=csv.reader(f)
    #row_n作为模板内容
    row_n=[]
    #设置修改内容相应内容的下标
    ti1=0
    de1=0
    sk1=0
    as1=0
    #设置模板内容对应下标
    ti2=0
    de2=0
    sk2=0
    as2=0
    #计数
    n=0
    for row in rows:
        #判断每一列代表什么
        for s in  range(len(row)):
            if "tem-name" in row[s]:
                ti1=s
            if "item-description" in row[s]:
                de1=s
            if "seller-sku" in row[s]:
                sk1=s
            if "asin1" in row[s]:
                as1=s
        if "tem-name" in str(row):
            continue

        #获取要写入模板的数据
        title=row[ti1]
        des=row[de1]
        sku=row[sk1]
        asin=row[as1]
        #标头部分直接读取写入
        if n<4:
            for i in range(3):
                for j in range(mysheet.max_column):
                    #获取标头的值
                    cel=mysheet.cell(i+1,j+1).value
                    #写入新实例 
                    sheet.cell(row=i+1,column= j+1,value=cel)
            #获取第四行内容作为正文模板 row_n=[]
            row_n=[]
            for j2 in range(mysheet.max_column):
                cel1=mysheet.cell(4,j2+1).value
                row_n.append(cel1)
            #通过第三行内容确定列所对应的内容
            for j3 in range(mysheet.max_column):
                cel3=mysheet.cell(3,j3+1).value
                if "item_sku" in cel3:
                    sk2=j3
                if "item_name" in cel3:
                    ti2=j3
                if "external_product_id" in cel3:
                    as2=j3
                if "product_description" in cel3:
                    de2=j3
            n=4
        #正文内容通过修改模板内容然后写入
        row_n[sk2]=sku.strip()
        row_n[ti2]=title.strip()
        row_n[as2]=asin.strip()
        row_n[de2]=des.strip()
        for j3 in range(len(row_n)):
            sheet.cell(row=n,column= j3+1,value=row_n[j3])
        n=n+1
    work.save(r"C:\Users\ZG\Downloads\211109待修改\New_MUG.xlsm")

